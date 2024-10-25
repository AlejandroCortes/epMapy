import numpy as np
import pandas as pd
import os
import glob
# import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt


pd.set_option('display.max_columns', 5)
pd.set_option('display.max_rows', 500)
#from openpyxl.styles import Font, Border, Side, Alignment, Protection, PatternFill, Color

#Define the file paths
date_path = 'C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\Dates\\'
traces_path = 'C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\Traces\\'
elements_to_drop = ['Pb_ppm_m207',
                   'Pb_ppm_m208',
                   'U_ppm_m235',
                   ]
standard_file = 'C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\StandardData.csv'

test_file = 'C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\BatchLog.csv'

# batch_df = pd.read_csv(test_file, sep = ',')
# batch_df.columns = batch_df.columns.str.replace(' ', '_') 
# print(batch_df.Sample_Name)

# samples = [batch_df.Sample_Name.loc[i] for i in range (len(batch_df.Sample_Name))]
# print(samples)

stds = pd.read_csv('StandardData.csv', sep=',', dtype=str, header=0)
std_list = stds.columns.tolist()
int_std = 153000

frequency = 8
spot_size = 25
fluence = 2.5


def compile_files(file_path):
    txt_files = glob.glob(file_path + '/*.txt')
    df_list = (pd.read_csv(file, sep = '\t', header = 0) for file in txt_files)
    compiled_df = pd.concat(df_list, ignore_index = True)
    compiled_df.columns.values[0] = 'Sample'
    return compiled_df

def filter_traces(traces_path):
    #extract header lists for conc, 2se and LOD
    conc_head = [ele for ele in list(compile_files(traces_path)) if ele[-1].isdigit()]
    std_err_head = [ele for ele in list(compile_files(traces_path)) if ele.endswith('2SE')]
    lod_head = [ele for ele in list(compile_files(traces_path)) if ele.endswith('LOD')]
    #Remove cps of intstd columns
    del std_err_head[:2]
    del lod_head[:2]
    #create single trace element dataframe in correct format
    traces = pd.concat([compile_files(traces_path).filter(items = conc_head),
                        compile_files(traces_path).filter(items = std_err_head),
                        compile_files(traces_path).filter(items = lod_head)], axis = 1)
    #Replace 'Below LOD' with < value
    for ele in conc_head:
        for samp in range(len(traces.index)):
            if traces.loc[samp, ele] == 'Below LOD':
                traces.loc[samp, ele] = '<' + str(traces.loc[samp,ele + '_LOD'])
                traces.loc[samp, ele + '_Int2SE'] = '<' + str(traces.loc[samp,ele + '_LOD'])
    return traces, conc_head 

def drop_elements(df, elements_to_drop):
    traces = df.drop(elements_to_drop + [element + '_LOD' for element in elements_to_drop] + [element + '_Int2SE' for element in elements_to_drop], axis=1)
    return traces

def order_dates(date_path):



    dates_df =  compile_files(date_path)[['Sample', 'Comments', 'Total points',  
             'Final207_235', 'Final207_235_Prop2SE', 'Final206_238', 'Final206_238_Prop2SE',
             'ErrorCorrelation_6_38vs7_35','Final207_206','Final207_206_Prop2SE', 'FinalAge207_235', 'FinalAge207_235_Prop2SE', 
             'FinalAge206_238', 'FinalAge206_238_Prop2SE','FinalAge207_206','FinalAge207_206_Prop2SE']]
    
    dates_df = dates_df.rename(columns = {'Total points' : 'Number of slices', 
                               'Final207_235' : '207Pb/235U',
                                'Final207_235_Prop2SE':'207Pb/235U_2s.',
                               'Final206_238' : '206Pb/238U', 
                               '206Pb/238U_2s.' : 'Final206_238_Prop2SE',
                                'ErrorCorrelation_6_38vs7_35' : 'Corr. coeff.',
                                'Final207_206' : '207Pb/206Pb',
                                'Final207_206_Prop2SE':'207Pb/206Pb_2s.',
                                'FinalAge207_235' : 'Date_207Pb/235U (Ma)', 
                                'FinalAge207_235_Prop2SE' : 'Date_207Pb/235U_2s. (Ma)', 
                                'FinalAge206_238' : 'Date_206Pb/238U (Ma)', 
                                'FinalAge206_238_Prop2SE' : 'Date_206Pb/238U_2s. (Ma)',
                                'FinalAge207_206' : 'Date_207Pb/206Pb (Ma)',
                                'FinalAge207_206_Prop2SE' : 'Date_207Pb/206Pb_2s. (Ma)'})
    #Insert extra columns
    dates_df.insert(2, 'Fluence', fluence)
    dates_df.insert(2, 'Frequency (Hz)', frequency)
    dates_df.insert(2, 'Spot Size (um)', spot_size)
    

    #Rename sample based on comments ----- Might need to come back to this
    dates_df['Sample'] = dates_df['Comments'].str.split('_').str[0]


    

    return dates_df

#combine dates+traces
def combine_data(dates_df, traces_df):
    # file = 'C:\\Users\\buret\\OneDrive\\Documents\\Python Scripts\\Geochron_Compile\\test.xlsx'
    combined_data = pd.concat([dates_df,traces_df],axis=1)
    #filter standards
    run_standards = {}

    for std_name in std_list:
        mask = combined_data['Sample'].isin(stds[std_name])
        #std_name is the standard name so need to filter empty dataframes
        raw_stds= combined_data[mask]
        if raw_stds.empty == False:
            run_standards[std_name] = raw_stds
            combined_data = combined_data[~mask]

    #Get the list of standards that were run from the dictionary keys
    
    # print(combined_data)
    #Remove the NaNs if all columns are NaN
    combined_data = combined_data.dropna(thresh= 4)
    combined_data.insert(6, 'Date Interpretation', '')
    combined_data.insert(21, 'Si29_ppm', int_std)

    #Get list of sample Names
    sample_names = combined_data['Sample'].unique()
    
    return combined_data, run_standards



##############~~~~~~~XLSX stuff~~~~~~~#######################
def write_xlsx_files(dfs):
   
    combined_data, run_standards = dfs
    run_standards_list = list(run_standards.keys())
    
    wb = Workbook()

    ws = wb.active
    ws.title = 'Compiled Data'
    #populate Compiled Data sheet
    [wb['Compiled Data'].append(r) for r in dataframe_to_rows(combined_data, index=False, header=True)]

    #Create new sheet names based on standard names and populate them based on dictionary
    [wb.create_sheet(standard_name) for standard_name in run_standards_list]
    [[wb[standard_name].append(r) for r in dataframe_to_rows(run_standards[standard_name], index=False, header=True)] 
     for standard_name in run_standards_list]

#Freeze all panes
    for sheet in wb.sheetnames:
        wb[sheet].freeze_panes = 'B1'
        

    wb.save('C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\VT_apatite.xlsx')

    return combined_data, run_standards

#Write excel sheets
traces, elements = filter_traces(traces_path)##########################################################################################

write_xlsx_files(combine_data(order_dates(date_path), drop_elements(traces, elements_to_drop))) #######################################

# anal_data, _ = combine_data(order_dates(date_path), drop_elements(traces, elements_to_drop))

# def analyse_data(anal_data, elements, elements_to_drop):
#     sample_names = anal_data['Sample'].unique()
#     print(elements)
#     x = anal_data['Hf_ppm_m178'].apply(pd.to_numeric, errors = 'coerce')
#     y = anal_data['Zr_ppm_m91'].apply(pd.to_numeric, errors = 'coerce')

#     plt.scatter(x,y)
#     plt.show()


# analyse_data(anal_data, elements, elements_to_drop)



#(combine_data(order_dates(date_path), drop_elements(filter_traces(traces_path), elements_to_drop)))






#Filter out the standards
#def filter_standards(combined_data):


#print(combine_data(order_dates(date_path), drop_elements(filter_traces(traces_path), elements_to_drop)))

# print(order_dates(date_path))
# print(drop_elements(filter_traces(traces_path), elements_to_drop))



#compiled_df.to_csv('C:\\Users\\yanb\\Documents\\Python Code\\Geochron_Compile\\comp_test.csv', encoding = 'utf-8')


# names = [
#     ['SPtab', 'sp', 'man'],
#     ['SBtab', 'sb', 'man'],
#     ['SDtab', 'sd', 'man']
# ]

# def create_dfs(names):
#     dfs = {}
#     for x in names:
#         dfs[x[1]] = pd.DataFrame()
#     return dfs
    
# dfs = create_dfs(names)

# print(dfs)
# print(list(dfs.keys())) 