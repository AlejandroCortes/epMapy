#importing libraries for standalone functions
import numpy as np
import pandas as pd
import re
import scipy.ndimage
import matplotlib.pyplot as plt
import tkinter as tk
import os
from openpyxl import load_workbook
import json
def show_intro(message):
    window = tk.Tk()  #Create the main window
    window.title("Please read this before continuing...") #Heading of the window
    window.geometry("400x500") #Window size
    label = tk.Label(window, text=message, padx=20, pady=20, wraplength=350) #Label widget for the message with word wrapping
    label.pack()
    button = tk.Button(window, text="Close", command=window.destroy) #Button widget to close the window
    button.pack(pady=10)
    window.mainloop() #Needed so that the code keeps running

def browse_win():
    filename = tk.filedialog.askopenfilename(initialdr = "/",
                                             title = "Select a file",
                                             filetype = (("Text files","*.DAT"),
                                                         ("Excel files","*.xlsx"),
                                                         ))
    tk.label_file_explorer.configure(text="file Opened: "+filename)

def clean_data(oxide_grids, total_grids):
    """
    Cleans the oxide grids based on the values of FeO, CaO, and Total oxides.
    
    Parameters:
    oxide_grids (dict): Dictionary containing oxide grids with oxides and their values.
    total_grids (dict): Dictionary containing 'Total' values for each (NX, NY) pair.
    
    Returns:
    cleaned_grids (dict): Dictionary containing cleaned oxide grids with bad data set to 0.
    """
    cleaned_grids = {}  # Dictionary to hold cleaned oxide grids



    # Iterate through each oxide in the oxide grids
    for oxide, grid in oxide_grids.items():
        clean_grid = {}  # To store cleaned data for this specific oxide
        
        # Iterate over the grid to clean the values based on Total, FeO, and CaO
        for (NX, NY), data in grid.items():
            
            # Retrieve FeO and CaO values for the current (NX, NY)
            feo_value = oxide_grids.get('FeO', {}).get((NX, NY), {}).get('oxide_value', 0) # Default to 0 if not found
            cao_value = oxide_grids.get('CaO', {}).get((NX, NY), {}).get('oxide_value', 0)  # Default to 0 if not found
            total_value = total_grids.get((NX, NY), 0)  # Get Total value from total_grids for the same (NX, NY)
            # Cleaning condition based on Total, FeO, and CaO
            if total_value < 80 and (feo_value < 40 or cao_value < 50):
                clean_grid[(NX, NY)] = {'oxide_value': 0, 'X_coord': data['X_coord'], 'Y_coord': data['Y_coord']}
                total_grids[(NX, NY)] = 0
            elif total_value > 105:
                clean_grid[(NX, NY)] = {'oxide_value': 0, 'X_coord': data['X_coord'], 'Y_coord': data['Y_coord']}
                total_grids[(NX, NY)] = 0
            else:
                clean_grid[(NX, NY)] = {'oxide_value': data['oxide_value'], 'X_coord': data['X_coord'], 'Y_coord': data['Y_coord']}
        
        # Store the cleaned grid for this oxide
        cleaned_grids[oxide] = clean_grid
    
    return cleaned_grids, total_grids


#This function takes major oxide compositions and normalise them to 100%
def to_anhydrous(oxide_grids, total_grids):
    anhydrous_grids = {}
    for oxide, grid in oxide_grids.items():
        anhydrous_grid = {}
        for (NX,NY), data in grid.items():
            total_value = total_grids.get((NX,NY),0)
            if total_value > 0 :
                anhydrous_grid[(NX, NY)] = {'oxide_value': 100*data['oxide_value']/total_value, 'X_coord': data['X_coord'], 'Y_coord': data['Y_coord']}
                total_grids[(NX, NY)] = 100
            else:
                anhydrous_grid[(NX, NY)] = {'oxide_value': 0, 'X_coord': data['X_coord'], 'Y_coord': data['Y_coord']}
        anhydrous_grids[oxide] = anhydrous_grid
    return anhydrous_grids, total_grids

def k_na(oxide_grids, total_grids):
    k_na_grid = {}

    # Directly access K2O and Na2O grids
    k2o_grid = oxide_grids.get('K2O', {})
    na2o_grid = oxide_grids.get('Na2O', {})

    # Iterate through the (NX, NY) pairs in the K2O grid
    for (NX, NY), data in k2o_grid.items():
        # Retrieve the K2O and Na2O values for the current (NX, NY)
        k2o_value = k2o_grid.get((NX, NY), {}).get('oxide_value', 0)
        na2o_value = na2o_grid.get((NX, NY), {}).get('oxide_value', 0)
        total_value = total_grids.get((NX, NY), 0)  # Get Total value from total_grids for the same (NX, NY)

        # If total_value is greater than 0, calculate the k_na value
        if total_value > 0:
            k_na_grid[(NX, NY)] = {
                'oxide_value': 1.119043665947202*k2o_value / na2o_value,
                'X_coord': data['X_coord'],
                'Y_coord': data['Y_coord']
            }
        else:
            # If the total_value is not valid, set oxide_value to 0
            k_na_grid[(NX, NY)] = {
                'oxide_value': 0,
                'X_coord': data['X_coord'],
                'Y_coord': data['Y_coord']
            }
    
    return k_na_grid

def load_json(): # function to load elements_weights and oxides_weights from JSON files and return two dictionaries
    with open('elements_weights.json', 'r') as f: #dictionary with elements and their atomic weights
        elements_weights = json.load(f)  
    with open('oxides_weights.json', 'r') as f: # dictionary with oxides and their molecular weights
        oxides_weights = json.load(f)
    return elements_weights, oxides_weights



#This function takes major oxide compositions in wt.% and converts them to mol
def to_mol(anhydrous_grids, total_grids, oxides_weights):
    mol_grids = {}
    for oxide, grid in anhydrous_grids.items():
        mol_grid = {}
        mw_oxide = oxides_weights[oxide]['molecular_weight']
        print(oxide, mw_oxide)

        for (NX, NY), data in grid.items():
            total_value = total_grids.get((NX, NY), 0)  # Get Total value from total_grids for the same (NX, NY)
            if total_value > 0:
                mol_grid[(NX, NY)] = {
                    'oxide_value': data['oxide_value']/mw_oxide,
                    'X_coord': data['X_coord'],
                    'Y_coord': data['Y_coord']
                    }
            else:
                mol_grid[(NX, NY)] = {
                    'oxide_value': 0,
                    'X_coord': data['X_coord'],
                    'Y_coord': data['Y_coord']
                    }
        mol_grids[oxide] = mol_grid
    return mol_grids, total_grids

def to_cat(mol_grids, total_grids, oxides_weights):
    cat_grids = {}
    for oxide, grid in mol_grids.items():
        cat_grid = {}
        cat_oxide = oxides_weights[oxide]['non_oxygen_atoms']
        print(oxide, cat_oxide)

        for (NX, NY), data in grid.items():
            total_value = total_grids.get((NX, NY), 0)  # Get Total value from total_grids for the same (NX, NY)
            if total_value > 0:
                cat_grid[(NX, NY)] = {
                    'oxide_value': data['oxide_value']*cat_oxide,
                    'X_coord': data['X_coord'],
                    'Y_coord': data['Y_coord']
                    }
            else:
                cat_grid[(NX, NY)] = {
                    'oxide_value': 0,
                    'X_coord': data['X_coord'],
                    'Y_coord': data['Y_coord']
                    }
        cat_grids[oxide] = cat_grid
    return cat_grids, total_grids

def m_f(cat_grids, total_grids):
    mf_grid = {}

    # Directly access K2O and Na2O grids
    k2oc_grid = cat_grids.get('K2O', {})
    na2oc_grid = cat_grids.get('Na2O', {})
    caoc_grid = cat_grids.get('CaO', {})
    sio2c_grid = cat_grids.get('SiO2', {})
    al2o3c_grid = cat_grids.get('Al2O3', {})

    # Iterate through the (NX, NY) pairs in the K2O grid
    for (NX, NY), data in k2oc_grid.items():
        # Retrieve the K2O and Na2O values for the current (NX, NY)
        k2oc_value = k2oc_grid.get((NX, NY), {}).get('oxide_value', 0)
        na2oc_value = na2oc_grid.get((NX, NY), {}).get('oxide_value', 0)
        caoc_value = caoc_grid.get((NX, NY), {}).get('oxide_value', 0)
        sio2c_value = sio2c_grid.get((NX, NY), {}).get('oxide_value', 0)
        al2o3c_value = al2o3c_grid.get((NX, NY), {}).get('oxide_value', 0)
        total_value = total_grids.get((NX, NY), 0)  # Get Total value from total_grids for the same (NX, NY)
        total_cat = k2oc_value+na2oc_value+caoc_value+sio2c_value+al2o3c_value
        print(type(total_cat),total_cat)

        # If total_value is greater than 0, calculate the k_na value
        if total_value > 0:
            mf_grid[(NX, NY)] = {
                'oxide_value': total_cat*(na2oc_value+k2oc_value+(2*caoc_value))/(sio2c_value*al2o3c_value),
                'X_coord': data['X_coord'],
                'Y_coord': data['Y_coord']
            }
        else:
            # If the total_value is not valid, set oxide_value to 0
            mf_grid[(NX, NY)] = {
                'oxide_value': 0,
                'X_coord': data['X_coord'],
                'Y_coord': data['Y_coord']
            }
    
    return mf_grid, total_grids
#This function takes major oxide compositions in mol and converts them to cation fractions

def add_fe2o3(X, oxide_column_indices):
    Fe2_Fetot = 0.7  # Ratio Fe2+/Fe3+ for NNO
    X2 = np.copy(X)  # Copy the input to avoid overwriting
    oxide_column_indices_fe = oxide_column_indices.copy()

    # Check if 'Fe2O3' column already exists in oxide_column_indices
    if "Fe2O3" not in oxide_column_indices_fe:
        # Calculate Fe2O3 from FeOtot and Fe2+/Fetot ratio
        feo_index = oxide_column_indices_fe.get('FeO', None)
        Fe2O3 = X[feo_index,:] * (1 - Fe2_Fetot) * 1.11134
        
        # Add the Fe2O3 data to the array
        X2 = np.vstack([X2, Fe2O3])  # Add Fe2O3 as a new row
        
        # Update FeO with the Fe2+/Fetot ratio
        X2[feo_index,:] = X2[feo_index,:] * Fe2_Fetot
    
        # Update the oxide_column_indices to include Fe2O3
        oxide_column_indices_fe["Fe2O3"] = len(oxide_column_indices_fe)  # Add Fe2O3 at the next available index
    else:
        # If Fe2O3 already exists, just return the original array without changes
        X2 = np.copy(X)  # No changes to the input array
        print("Fe2O3 column already exists, no update performed.")
    
    return X2, oxide_column_indices_fe

def norm_calc(X, oxide_column_indices):
    Z = np.copy(X)  # Copy the input to avoid overwriting
    a = len(Z[0,:])  # Number of pixels
    oots = np.zeros((a, 23))  # Array to store calculated values
    fps=np.zeros((a,8)) #First pass silica
    sps=np.zeros((a,11)) #Second pass silica
    fs=np.zeros((a,7)) #final silica
    n_wt=np.zeros((a,16)) #recalculate to normative mineral in wt%

    # Define the oxides we're interested in and map them to indices in oxide_column_indices
    oxide_names = ['SiO2', 'TiO2', 'Al2O3', 'Fe2O3', 'FeO', 'MnO', 'MgO', 'CaO', 'Na2O', 'K2O','P2O5', 'Total']
    oots_names = [
        "ap_Ca", "ap_P", "ilm_Fe_II", "Ilm_Ti", "or_lc_Al", "or_lc_K", "ab_ne_Al", "ab_ne_Na", 
        "ac_Na", "ac_FeIII", "ns_Na", "mt_Fe_II", "mt_Fe_III", "an_Al_prelim", "an_Al_final", 
        "C_Al", "an_Ca", "di_Ca", "di_Mg", "di_FeII", "ol_hy_Fe_II", "ol_hy_Mg", "Mg_num"
        ]
    fps_names = ["lc", "ne", "ac", "ns", "an", "di", "ol", "residual_si"]
    sps_names = [
        "or_K", "lc_K", "residual_si2", "ab_Na", "ne_Na","residual_si3", "prelim_ol_Mg_Fe", "prelim_hy_Mg_Fe", 
        "actual_ol_Mg_Fe", "actual_hy_Mg_Fe", "residual_siQ"
        ]
    fs_names = ["or", "lc", "ab", "ne", "ol", "hy", "Q"]
    nwt_names = ["Q", "or", "lc", "ab", "ne", "an", "C", "ac", "ns", "di", "ol", "hy", "mt", "ilm", "ap","total"]

    mw = {
        "SiO2": 60.080, "TiO2": 79.870, "Cr2O3": 152.000, "Al2O3": 101.960, "FeO": 71.840, 
        "MnO": 70.940, "NiO": 74.690, "MgO": 40.300, "CaO": 56.080, "Na2O": 61.980, 
        "K2O": 94.200, "P2O5": 283.880, "F": 18.998, "Cl": 35.450, "SO3": 80.060, 
        "BaO": 153.330, "SrO": 103.620, "La2O3": 325.81, "Ce2O3": 328.24, "PbO":223.2, "Fe2O3": 141.94
    }

    oxide_indices = {oxide: oxide_column_indices.get(oxide, None) for oxide in oxide_names}
    oots_index = {name: index for index, name in enumerate(oots_names)}
    fps_index = {name: index for index, name in enumerate(fps_names)}
    sps_index = {name: index for index, name in enumerate(sps_names)}
    fs_index = {name: index for index, name in enumerate(fs_names)}
    nwt_index = {name: index for index, name in enumerate(nwt_names)}

    condA = Z[oxide_indices['Total'],:] > 0

    oots[:, oots_index["ap_Ca"]] = Z[oxide_indices['P2O5']] * 10 / 3  # ap Ca
    oots[:, oots_index["ap_P"]] = Z[oxide_indices['P2O5']]  # ap P
    oots[:, oots_index["ilm_Fe_II"]] = Z[oxide_indices['TiO2']]  # ilm Fe(II)
    oots[:, oots_index["Ilm_Ti"]] = Z[oxide_indices['TiO2']]  # ilm Ti
    oots[:, oots_index["or_lc_Al"]] = Z[oxide_indices['K2O']]  # or/lc Al
    oots[:, oots_index["or_lc_K"]] = Z[oxide_indices['K2O']]  # or/lc K

    oots[:, oots_index["ab_ne_Al"]] = np.where(condA & (Z[oxide_indices['Na2O']] > (Z[oxide_indices['Al2O3']] - oots[:, oots_index["or_lc_Al"]])),
                                               Z[oxide_indices['Al2O3']] - oots[:, oots_index["or_lc_Al"]],
                                               Z[oxide_indices['Na2O']]) 
    oots[:, oots_index["ab_ne_Na"]] = np.where(condA,
                                                oots[:, oots_index["ab_ne_Al"]],0)
    oots[:, oots_index["ac_FeIII"]] = np.where(condA & ((Z[oxide_indices['Na2O']] - oots[:, oots_index["ab_ne_Na"]]) < Z[oxide_indices['Fe2O3']]),
                                               Z[oxide_indices['Na2O']] - oots[:, oots_index["ab_ne_Na"]],
                                               Z[oxide_indices['Fe2O3']])
    oots[:, oots_index["ac_Na"]] = np.where(condA, oots[:, oots_index["ac_FeIII"]],0)  # ac Na
    oots[:, oots_index["ns_Na"]] = np.where(condA, Z[oxide_indices['Na2O']] - oots[:, oots_index["ab_ne_Na"]] - oots[:, oots_index["ac_Na"]],0)  # ns Na
    oots[:, oots_index["mt_Fe_III"]] = np.where(condA, Z[oxide_indices['Fe2O3']] - oots[:, oots_index["ac_FeIII"]],0)  # mt Fe(III)
    oots[:, oots_index["mt_Fe_II"]] = np.where(condA, oots[:, oots_index["mt_Fe_III"]],0)
    oots[:, oots_index["an_Al_prelim"]] = np.where(condA & (oots[:, oots_index["ac_Na"]] > 0.000001),
                                                   0,
                                                   Z[oxide_indices['Al2O3']] - oots[:, oots_index["or_lc_Al"]] - oots[:, oots_index["ab_ne_Al"]])
    oots[:, oots_index["an_Al_final"]] = np.where(condA & (oots[:, oots_index["an_Al_prelim"]] > (Z[oxide_indices['CaO']] - oots[:, oots_index["ap_Ca"]])),
                                                  Z[oxide_indices['CaO']] - oots[:, oots_index["ap_Ca"]],
                                                  oots[:, oots_index["an_Al_prelim"]])                                               
    oots[:, oots_index["C_Al"]] = np.where(condA &(oots[:, oots_index["an_Al_prelim"]] == oots[:, oots_index["an_Al_final"]]),
                                           0,
                                           oots[:, oots_index["an_Al_prelim"]] - oots[:, oots_index["an_Al_final"]])
    oots[:, oots_index["an_Ca"]] = np.where(condA, oots[:, oots_index["an_Al_final"]],0)  # an Ca
    oots[:, oots_index["di_Ca"]] = np.where(condA, Z[oxide_indices['CaO']] - oots[:, oots_index["ap_Ca"]] - oots[:, oots_index["an_Ca"]],0)  # di Ca
    oots[:, oots_index["Mg_num"]] = np.where(condA, Z[oxide_indices['MgO']] / (Z[oxide_indices['MgO']] + Z[oxide_indices['FeO']] + Z[oxide_indices['MnO']] - oots[:, oots_index["ilm_Fe_II"]] - oots[:, oots_index["mt_Fe_II"]]),0)  # XMg
    oots[:, oots_index["di_Mg"]] = np.where(condA, oots[:, oots_index["di_Ca"]] * oots[:, oots_index["Mg_num"]],0)  # di Mg
    oots[:, oots_index["di_FeII"]] = np.where(condA, oots[:, oots_index["di_Ca"]] * (1 - oots[:, oots_index["Mg_num"]]),0)  # di FeII
    oots[:, oots_index["ol_hy_Fe_II"]] = np.where(condA, Z[oxide_indices['FeO']] + Z[oxide_indices['MnO']] - oots[:, oots_index["ilm_Fe_II"]] - 
                                                  oots[:, oots_index["mt_Fe_II"]] - oots[:, oots_index["di_FeII"]],0)  # ol/hy Fe(II)
    oots[:, oots_index["ol_hy_Mg"]] = np.where(condA,Z[oxide_indices['MgO']] - oots[:, oots_index["di_Mg"]],0)  # ol/hy Mg
    
    fps[:, fps_index["lc"]]   = np.where(condA,4*oots[:, oots_index["or_lc_K"]],0)                               #lc
    fps[:, fps_index["ne"]]   = np.where(condA,2*oots[:, oots_index["ab_ne_Na"]],0)                               #ne
    fps[:, fps_index["ac"]]   = np.where(condA,4*oots[:, oots_index["ac_Na"]],0)                                        #ac
    fps[:, fps_index["ns"]]   = np.where(condA,oots[:, oots_index["ns_Na"]],0)                                         #ns
    fps[:, fps_index["an"]]   = np.where(condA,2*oots[:, oots_index["an_Ca"]],0)                                       #an
    fps[:, fps_index["di"]]   = np.where(condA,2*oots[:, oots_index["di_Ca"]],0)                                       #di
    fps[:, fps_index["ol"]]   = np.where(condA,(oots[:, oots_index["ol_hy_Fe_II"]]+ oots[:, oots_index["ol_hy_Mg"]])/2,0)  
    
    fps_columns = [index for col_name, index in fps_index.items() if col_name != "residual_si"]
    sum_of_fps = np.where(condA,np.sum(fps[:, fps_columns], axis=1),0)  # sum all columns except "residual_si"
    fps[:, fps_index["residual_si"]]   = np.where(condA,Z[oxide_indices["SiO2"]]-sum_of_fps,0) #residual Si
    
    sps[:, sps_index["or_K"]] = np.where(condA & ((oots[:, oots_index["or_lc_K"]]*2) > fps[:, fps_index["residual_si"]]),
                                        (fps[:, fps_index["lc"]]+(fps[:,fps_index["residual_si"]]-(oots[:, oots_index["or_lc_K"]]*2)))/2, #or K
                                         oots[:, oots_index["or_lc_K"]])                             #or K 
    sps[:, sps_index["lc_K"]]  = np.where(condA,oots[:, oots_index["or_lc_K"]]-sps[:, sps_index["or_K"]],0)                         #lc K
    sps[:, sps_index["residual_si2"]]  = np.where(condA,fps[:, fps_index["residual_si"]]-(2*sps[:, sps_index["or_K"]]),0)                      #residual Si
    sps[:, sps_index["ab_Na"]] = np.where(condA & ((oots[:, oots_index["ab_ne_Na"]]*4) > sps[:, sps_index["residual_si2"]]),
                                          (sps[:, sps_index["residual_si2"]]+fps[:, fps_index["ne"]]-(2*oots[:, oots_index["ab_ne_Na"]]))/4,
                                          oots[:, oots_index["ab_ne_Na"]])
    sps[:, sps_index["ne_Na"]]  = np.where(condA,oots[:, oots_index["ab_ne_Na"]]-sps[:, sps_index["ab_Na"]],0)                         #ne Na
    sps[:, sps_index["residual_si3"]]  = np.where(condA,sps[:, sps_index["residual_si2"]]-(4*sps[:, sps_index["ab_Na"]]),0)                      #residual Si
    sps[:, sps_index["prelim_ol_Mg_Fe"]] = np.where(condA & (sps[:, sps_index["ne_Na"]] > 0.0000001),
                                                    oots[:, oots_index["ol_hy_Fe_II"]]+oots[:, oots_index["ol_hy_Mg"]],
                                                    oots[:, oots_index["ol_hy_Fe_II"]]+oots[:, oots_index["ol_hy_Mg"]]-(2*sps[:, sps_index["prelim_ol_Mg_Fe"]]))
    sps[:, sps_index["prelim_hy_Mg_Fe"]]  = np.where(condA,oots[:, oots_index["ol_hy_Fe_II"]]+oots[:, oots_index["ol_hy_Mg"]]-sps[:, sps_index["prelim_ol_Mg_Fe"]],0)             #prelim hy Mg+Fe
    condB = sps[:, sps_index["prelim_ol_Mg_Fe"]] < 0
    sps[:, sps_index["actual_ol_Mg_Fe"]] = np.where(condA & condB,
                                                    0,
                                                    sps[:, sps_index["prelim_ol_Mg_Fe"]])
    sps[:, sps_index["actual_hy_Mg_Fe"]] = np.where(condA & condB,
                                                    oots[:, oots_index["ol_hy_Fe_II"]]+oots[:, oots_index["ol_hy_Mg"]],
                                                    sps[:, sps_index["prelim_hy_Mg_Fe"]])
    sps[:, sps_index["residual_siQ"]] = np.where(condA & condB,
                                                 sps[:, sps_index["residual_si3"]]-((oots[:, oots_index["ol_hy_Fe_II"]]+oots[:, oots_index["ol_hy_Mg"]])/2),
                                                 0)
                                                 
    fs[:, fs_index["or"]]  = np.where(condA,6*sps[:, sps_index["or_K"]],0)   #or
    fs[:, fs_index["lc"]]  = np.where(condA,4*sps[:, sps_index["lc_K"]],0)   #lc
    fs[:, fs_index["ab"]]  = np.where(condA,6*sps[:, sps_index["ab_Na"]] ,0)  #ab
    fs[:, fs_index["ne"]]  = np.where(condA,2*sps[:, sps_index["ne_Na"]] ,0)  #ne
    fs[:, fs_index["ol"]]  = np.where(condA,0.5*sps[:, sps_index["actual_ol_Mg_Fe"]],0) #ol
    fs[:, fs_index["hy"]]  = np.where(condA,sps[:, sps_index["actual_hy_Mg_Fe"]] ,0)    #hy
    fs[:, fs_index["Q"]]  = np.where(condA,sps[:,sps_index["residual_siQ"]],0)    #Q
           
    n_wt[:, nwt_index["Q"]]  = np.where(condA,fs[:, fs_index["Q"]]*60.08,0)                                                 #Q
    n_wt[:, nwt_index["or"]]  = np.where(condA,sps[:, sps_index["or_K"]]*556.64,0)                                               #or
    n_wt[:, nwt_index["lc"]]  = np.where(condA,sps[:, sps_index["lc_K"]]*436.48,0)                                               #lc
    n_wt[:, nwt_index["ab"]]  = np.where(condA,sps[:, sps_index["ab_Na"]]*524.42,0)                                               #ab
    n_wt[:, nwt_index["ne"]]  = np.where(condA,sps[:, sps_index["ne_Na"]]*284.1,0)                                                #ne
    n_wt[:, nwt_index["an"]]  = np.where(condA,oots[:, oots_index["an_Ca"]]*278.2 ,0)                                             #an
    n_wt[:, nwt_index["C"]]  = np.where(condA,oots[:, oots_index["C_Al"]]*101.96 ,0)                                            #C
    n_wt[:, nwt_index["ac"]]  = np.where(condA,oots[:, oots_index["ac_Na"]]*462  ,0)                                               #ac
    n_wt[:, nwt_index["ns"]]  = np.where(condA,oots[:, oots_index["ns_Na"]]*122.06 ,0 )                                           #ns
    n_wt[:, nwt_index["di"]]  = np.where(condA,(oots[:, oots_index["di_Ca"]]*116.16)+(oots[:, oots_index["di_Mg"]]*100.38)+(oots[:, oots_index["di_FeII"]]*131.39),0)   #di
    n_wt[:, nwt_index["ol"]] = np.where(condA,(sps[:, sps_index["actual_ol_Mg_Fe"]]*70.34*oots[:, oots_index["Mg_num"]])+(sps[:, sps_index["actual_ol_Mg_Fe"]]*(1-oots[:, oots_index["Mg_num"]])*101.89),0)  #ol
    n_wt[:, nwt_index["hy"]] = np.where(condA,(sps[:, sps_index["actual_hy_Mg_Fe"]]*100.38*oots[:, oots_index["Mg_num"]])+(sps[:, sps_index["actual_hy_Mg_Fe"]]*(1-oots[:, oots_index["Mg_num"]])*131.93),0 )#hy
    n_wt[:, nwt_index["mt"]] = np.where(condA,oots[:, oots_index["mt_Fe_III"]]*231.55,0   )                                          #mt
    n_wt[:, nwt_index["ilm"]] = np.where(condA,oots[:, oots_index["Ilm_Ti"]]*151.72 ,0  )                                           #ilm
    n_wt[:, nwt_index["ap"]] = np.where(condA,oots[:, oots_index["ap_P"]]*328.87   ,0  )                                         #ap
    nwt_columns = [index for col_name, index in nwt_index.items() if col_name != "total"]                       # Exclude "total"
    sum_of_nwt = np.where(condA,np.sum(n_wt[:, nwt_columns], axis=1),0)                                                                    # sum all columns except "residual_si"
    n_wt[:, nwt_index["total"]] = np.where(condA,sum_of_nwt,0)

    return n_wt, nwt_index

###########################################################################################################################################################
# Function for binary profile extraction
###########################################################################################################################################################
# def extract_profile(grid, oxides, pixel_size, selected_oxide, sample_name, profile_index):
#     print("Select two points on the grid to define a traverse.")
    
#     # Create a figure to display the grid
#     plt.figure(figsize=(6, 6))
#     im = plt.imshow(grid[selected_oxide], cmap='viridis', vmin=0, vmax=np.max(grid[selected_oxide]))
#     plt.colorbar(im, label=f"{selected_oxide} wt.%")
#     plt.title(f"Select two points for {selected_oxide} profile")
#     plt.axis('off')  # Hide axes for cleaner view
    
#     # Ensure the plot is rendered first before selecting points
#     plt.show(block=False)  # This allows interaction while the plot stays open
#     points = plt.ginput(2)  # Allow the user to select two points on the plot
    
#     if len(points) < 2:
#         print("Traverse selection canceled.")
#         return

#     # Unpack the coordinates of the two points
#     (x1, y1), (x2, y2) = points
#     print(f"Selected points: ({x1:.2f}, {y1:.2f}) to ({x2:.2f}, {y2:.2f})")

#     # Create an array of x and y coordinates between the two points (traverse)
#     num_points = 80
#     x_coords = np.linspace(x1, x2, num_points)
#     y_coords = np.linspace(y1, y2, num_points)

#     # Debugging: Print coordinates to make sure they're being generated correctly
#     print(f"x_coords: {x_coords}")
#     print(f"y_coords: {y_coords}")

#     # Map the coordinates onto the grid to get the profile values (using scipy's map_coordinates)
#     profiles = {}
#     for oxide in oxides:
#         profile_values = scipy.ndimage.map_coordinates(grid[oxide], [y_coords, x_coords], order=1)
#         profiles[oxide] = profile_values
#     # Debugging: Check if the profile values are being calculated
#     for oxide, profile_values in profiles.items():
#         print(f"Profile values for {oxide}: {profile_values}")

#     # Get the scaling factors for X and Y coordinates from coord_data
#     # Convert pixel distances to micrometers
#     distance_pixels = np.sqrt((x_coords - x1) ** 2 + (y_coords - y1) ** 2)
#     distance_microns = distance_pixels * pixel_size
    
#     # Debugging: Check the calculated micrometer distances
#     print(f"Distance in micrometers: {distance_microns}")

#     # Create subplots for each oxide profile
#     num_oxides = len(oxides)
#     fig, axs = plt.subplots(1, num_oxides, figsize=(3 * num_oxides, 3))  # Create subplots horizontally

#     if num_oxides == 1:  # If only one subplot, make axs a list to handle consistently
#         axs = [axs]

#     for i, (oxide, profile_values) in enumerate(profiles.items()):
#         ax = axs[i]  # Get the corresponding subplot for each oxide
#         # Calculate error as 10% of the profile values
        
#         # Plot each oxide's profile with error bars
#         ax.scatter(distance_microns, profile_values, label=f"{oxide}", edgecolor='black', linewidths=0.5)

#         ax.set_title(f"Profile of {oxide}")
#         ax.set_xlabel(r"Distance ($\mu$m)")
#         ax.set_ylabel("Concentration (wt.%)")
#         ax.legend()  # Add a legend to each subplot

#     plt.tight_layout()

#     # Save the plot as a PDF file
#     pdf_filename = f"{sample_name}_profiles.pdf"
#     plt.savefig(pdf_filename, dpi=600, transparent=True, bbox_inches='tight')
#     print(f"Plot saved as {pdf_filename}")

#     # Show the plot
#     plt.show()

#     # --- Save the data to Excel ---
#     # Prepare a dictionary to store the traverse data
#     data_dict = {'Distance (µm)': distance_microns}
    
#     # Add oxide profile values to the dictionary
#     for oxide, profile_values in profiles.items():
#         data_dict[f'{oxide} (wt.%)'] = profile_values
    
#     # Convert dictionary to pandas DataFrame
#     df = pd.DataFrame(data_dict)
    
#     # Save the DataFrame to an Excel file
#     excel_filename = f"{sample_name}_profiles_data.xlsx"
#     with pd.ExcelWriter(excel_filename, engine='xlsxwriter') as writer:
#         df.to_excel(writer, index=False, sheet_name='Traverse Data')
#     print(f"Data saved to {excel_filename}")


import numpy as np
import matplotlib.pyplot as plt
import scipy.ndimage

# def extract_profile(grid, oxides, pixel_size, selected_oxide, sample_name, profile_index):
#     print("Select two points on the grid to define a traverse.")
    
#     # Convert the dictionary of (NX, NY) pairs into a 2D numpy array for the selected oxide
#     # Get the grid size based on the NX, NY coordinates
#     oxides_data = grid[selected_oxide]
#     max_x = max([coord[0] for coord in oxides_data.keys()])
#     max_y = max([coord[1] for coord in oxides_data.keys()])
    
#     # Create a 2D array for the selected oxide, filled with NaN values initially
#     oxide_array = np.full((max_x + 1, max_y + 1), np.nan)
    
#     # Fill the array with the actual oxide values from the (NX, NY) pairs
#     for (NX, NY), oxide_data in oxides_data.items():
#     # Extract the oxide_value from the dictionary
#         oxide_value = oxide_data.get('oxide_value', np.nan)  # Default to np.nan if not found
#         oxide_array[NX, NY] = oxide_value

#     # Plot the grid with the selected oxide using imshow()
#     plt.figure(figsize=(6, 6))
#     im = plt.imshow(oxide_array, cmap='viridis', vmin=0, vmax=np.nanmax(oxide_array))  # Handle NaN values
#     plt.colorbar(im, label=f"{selected_oxide} wt.%")
#     plt.title(f"Select two points for {selected_oxide} profile")
#     plt.axis('off')  # Hide axes for cleaner view
    
#     # Ensure the plot is rendered first before selecting points
#     plt.show(block=False)  # This allows interaction while the plot stays open
#     points = plt.ginput(2)  # Allow the user to select two points on the plot
    
#     if len(points) < 2:
#         print("Traverse selection canceled.")
#         return

#     # Unpack the coordinates of the two points
#     (x1, y1), (x2, y2) = points
#     print(f"Selected points: ({x1:.2f}, {y1:.2f}) to ({x2:.2f}, {y2:.2f})")

#     # Ensure the distance between the two points is at least 1.5 times the pixel size
#     distance_pixels = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
#     min_distance = 1.5 * pixel_size  # Minimum distance between points
#     if distance_pixels < min_distance:
#         print(f"Error: The distance between points is too small. Please select a wider range (at least {min_distance} pixels).")
#         return
    
#     # Create an array of x and y coordinates between the two points (traverse)
#     num_points = max(int(distance_pixels // min_distance), 2)  # Ensure at least 2 points
#     x_coords = np.linspace(x1, x2, num_points)
#     y_coords = np.linspace(y1, y2, num_points)

#     # Debugging: Print coordinates to make sure they're being generated correctly
#     print(f"x_coords: {x_coords}")
#     print(f"y_coords: {y_coords}")

#     # Map the coordinates onto the grid to get the profile values (using scipy's map_coordinates)
#     profiles = {}
#     for oxide in oxides:
#         profile_values = scipy.ndimage.map_coordinates(oxide_array, [y_coords, x_coords], order=1)
#         profiles[oxide] = profile_values
#     # Debugging: Check if the profile values are being calculated
#     for oxide, profile_values in profiles.items():
#         print(f"Profile values for {oxide}: {profile_values}")

#     # Get the scaling factors for X and Y coordinates from coord_data
#     # Convert pixel distances to micrometers
#     distance_pixels = np.sqrt((x_coords - x1) ** 2 + (y_coords - y1) ** 2)
#     distance_microns = distance_pixels * pixel_size
    
#     # Debugging: Check the calculated micrometer distances
#     print(f"Distance in micrometers: {distance_microns}")

#     # Plot the image with the selected line (profile)
#     fig, ax = plt.subplots(figsize=(6, 6))
#     im = ax.imshow(oxide_array, cmap='viridis', vmin=0, vmax=np.nanmax(oxide_array))
#     ax.plot([x1, x2], [y1, y2], color='red', linewidth=2, label="Profile Line")  # Draw the profile line
#     ax.arrow(x1,y1,x2-x1,y2-y1,width=1,facecolor='black',edgecolor='white',length_includes_head=True)
#     ax.set_title(f"{selected_oxide} with Profile Line")
#     ax.axis('off')
#     ax.legend()

#     # Save the image with the profile line
#     profile_image_filename = f"{sample_name}_profile_{profile_index}.png"
#     plt.savefig(profile_image_filename, dpi=300, transparent=True, bbox_inches='tight')
#     print(f"Profile image saved as {profile_image_filename}")
#     plt.close()  # Close the plot to avoid excessive memory usage

#     # Prepare profile data
#     data_dict = {'Distance (µm)': distance_microns}
#     for oxide, profile_values in profiles.items():
#         data_dict[f'{oxide} (wt.%)'] = profile_values
#     df = pd.DataFrame(data_dict)

#     # Define the Excel filename
#     excel_filename = f"{sample_name}_profiles_data.xlsx"

#     # Check if the file exists
#     if not os.path.exists(excel_filename):
#         # Create an empty file if it doesn't exist
#         with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
#             # Write an empty dataframe or just create the sheet to initialize the file
#             pd.DataFrame().to_excel(writer, index=False, sheet_name="Sheet1")

#     # Create a new sheet name based on profile index
#     sheet_name = f"Profile_{profile_index}"

#     # Load the existing workbook to check if the sheet exists
#     book = load_workbook(excel_filename)

#     # If the sheet already exists, delete it
#     if sheet_name in book.sheetnames:
#         del book[sheet_name]

#     # Append data to the file (with the sheet being replaced if it exists)
#     with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a') as writer:
#         # Write the data to the specified sheet name
#         df.to_excel(writer, index=False, sheet_name=sheet_name)

#     print(f"Data saved to {excel_filename} in sheet {sheet_name}")
